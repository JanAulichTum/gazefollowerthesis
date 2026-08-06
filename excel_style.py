# -*- coding: utf-8 -*-
"""
Shared Excel styling for all study workbooks.

Keeps the data files presentable for the thesis appendix and pleasant to
work with: bold header on a TUM-blue fill, frozen header row, an
auto-filter, and sensible column widths. Styling never changes the data
itself.
"""

from __future__ import annotations

import logging

logger = logging.getLogger(__name__)

HEADER_FILL = "0065BD"      # TUM blue
HEADER_FONT_COLOR = "FFFFFF"
MAX_COL_WIDTH = 42
MIN_COL_WIDTH = 10
WIDTH_SAMPLE_ROWS = 50      # rows sampled for width estimation (speed)


def style_workbook(path: str) -> None:
    """Apply the house style to the first sheet of an .xlsx file.

    Safe to call on every write: on any error the file is left as
    written (data always wins over looks).
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = load_workbook(path)
        ws = wb.active
        if ws.max_row < 1 or ws.max_column < 1:
            return

        header_font = Font(bold=True, color=HEADER_FONT_COLOR)
        header_fill = PatternFill(
            start_color=HEADER_FILL, end_color=HEADER_FILL,
            fill_type="solid",
        )
        header_align = Alignment(vertical="center")

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

            # Column width: header length vs. a sample of the values
            longest = len(str(cell.value or ""))
            last = min(ws.max_row, 1 + WIDTH_SAMPLE_ROWS)
            for row in range(2, last + 1):
                v = ws.cell(row=row, column=col).value
                if v is not None:
                    longest = max(longest, len(str(v)))
            ws.column_dimensions[get_column_letter(col)].width = max(
                MIN_COL_WIDTH, min(longest + 2, MAX_COL_WIDTH))

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        wb.save(path)
    except Exception:  # noqa: BLE001 — cosmetics must never break data
        logger.exception("Workbook styling failed (data unaffected): %s",
                         path)
